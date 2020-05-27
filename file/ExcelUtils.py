#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import openpyxl
import xlrd
import xlwt
import datetime
from xlrd import xldate_as_tuple

# xlrd：对xls、xlsx、xlsm文件进行读操作–读操作效率较高，推荐
# xlwt：对xls文件进行写操作–写操作效率较高，但是不能执行xlsx文件
# openpyxl：对xlsx、xlsm文件进行读、写操作–xlsx写操作推荐使用


def read_excel(sheet):
    # 定义一个空列表
    sheet_data = []
    for i in range(1, sheet.nrows):
        # 定义一个空字典
        row_data = {}
        for j in range(sheet.ncols):
            # 获取单元格数据类型
            c_type = sheet.cell(i, j).ctype
            # 获取单元格数据
            c_cell = sheet.cell_value(i, j)
            if c_type == 2 and c_cell % 1 == 0:  # 如果是整形
                c_cell = int(c_cell)
            elif c_type == 3:
                # 转成datetime对象
                date = datetime.datetime(*xldate_as_tuple(c_cell, 0))
                c_cell = date.strftime('%Y/%d/%m %H:%M:%S')
            elif c_type == 4:
                c_cell = True if c_cell == 1 else False
            row_data[sheet.row_values(0)[j]] = c_cell
            # 循环每一个有效的单元格，将字段与值对应存储到字典中
            # 字典的key就是excel表中每列第一行的字段
            # sheet_data[self.keys[j]] = self.table.row_values(i)[j]
        # 再将字典追加到列表中
        sheet_data.append(row_data)

    # 返回从excel中获取到的数据：以列表存字典的形式返回
    return sheet_data


class ReadExcelUtils:

    def __init__(self, file_path):
        # 定义一个属性接收文件路径
        self.file_path = file_path

        # 使用xlrd模块打开excel表读取数据
        self.workbook = xlrd.open_workbook(self.file_path)

    def get_sheet_names(self):
        return self.workbook.sheet_names()

    def get_sheet(self, sheet_name):
        return self.workbook.sheet_by_name(sheet_name)


class WriteXLSExcelUtils:

    def __init__(self, file_path):
        # 定义一个属性接收文件路径
        self.file_path = file_path

        # 使用xlrd模块打开excel表读取数据
        self.data = xlrd.open_workbook(self.file_path)

    def get_sheet_names(self):
        return self.data.sheet_names()

    def get_sheet(self, sheet_name):
        return self.data.sheet_by_name(sheet_name)


def set_sheet_name(sheet, sheet_name):
    sheet.title = sheet_name


class WriteXLSXExcelUtils:

    def __init__(self, file_path, sheet_name, titles, contents):
        # 使用xlrd模块打开excel表读取数据
        self.workbook = openpyxl.Workbook()

        # 定义一个属性接收文件路径
        self.file_path = file_path

        # 定义一个属性接收sheet名称
        self.sheet_name = sheet_name

        # 定义一个属性接收sheet中的标题
        self.titles = titles

        # 定义一个属性接收sheet中的内容
        self.contents = contents

    def set_sheet_title(self, sheet):
        for col in range(len(self.titles)):
            c = col + 1
            sheet.cell(row=1, column=c).value = self.titles[col]

    def set_sheet_contents(self, sheet):
        for col in range(len(self.contents)):
            sheet.append(self.contents[col])

    def save(self):
        return self.workbook.save(self.file_path)


def main():
    # 读excel文件
    filename = "d:/test.xlsx"
    get_data = ReadExcelUtils(filename)
    sheet_names = get_data.get_sheet_names()
    for name in sheet_names:
        print(read_excel(get_data.get_sheet(name)))

    # 写内容到xlsx格式的excel文件中
    titles = ["a", "b", "c", "d", "e"]
    contents = [
                ["a1", "b1", "c1", "d1", "e1"],
                {"a": "a2", "b": "b2", "c": "c2", "d": "d2", "e": "e2"},
                {1: "a3", 3: "c3", 5: "哈哈"}
        ]
    excel = WriteXLSXExcelUtils("d:/test1.xlsx", "你好", titles, contents)
    # sheet = excel.workbook.create_sheet(excel.sheet_name)
    sheet = excel.workbook.active
    # sheet.sheet_properties.tabColor = "1072BA"
    sheet.title = excel.sheet_name
    excel.set_sheet_title(sheet)
    excel.set_sheet_contents(sheet)
    excel.save()


if __name__ == '__main__':
    main()
